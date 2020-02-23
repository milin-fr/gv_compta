import tkinter
from tkinter import Label, Button, Entry, Checkbutton, OptionMenu, Canvas, Frame
from tkinter import StringVar, IntVar
from tkinter.messagebox import showinfo
import openpyxl
from openpyxl import Workbook, load_workbook
import datetime

'''Below module allows us to interact with Windows files.'''
import os

'''below 3 lines allows script to check the directory where it is executed, so it knows where to crete the excel file. I copied the whole block from stack overflow'''
abspath = os.path.abspath(__file__)
current_directory = os.path.dirname(abspath)
os.chdir(current_directory)

row_counter = 0

dictionary_of_name_entrie = {}
dictionary_of_date_entrie = {}
dictionary_of_payment_entrie = {}
dictionary_of_last_date_labele = {}
dictionary_of_last_date_labele_value = {}
dictionary_of_last_payment_labele = {}
dictionary_of_last_payment_labele_value = {}
dictionary_of_payment_total_labeles = {}
dictionary_of_payment_total_labeles_value = {}

name_index_dictionary = {}
master_dictionary = {}

def get_files_in_script_directory():
    '''Get file names in directory'''
    file_names = []
    for root, dirs, files in os.walk(current_directory):
        for filename in files:
            file_names.append(filename)
    return file_names

def get_date_yyyy_mm_dd():
    return str(datetime.date.today())


def get_current_year():
    #current_month = datetime.datetime.now().strftime('%B') #extracts the name of the current month
    current_year = datetime.datetime.now().strftime('%Y') #extracts the current year
    return str(current_year)


def generate_the_excel_file_name_with_current_year_in_name():
    return (get_current_year() + " sandbox.xlsx")


def check_if_excel_file_is_there():
    return generate_the_excel_file_name_with_current_year_in_name() in get_files_in_script_directory()


def create_excel_file_if_it_was_not_there():
    if not check_if_excel_file_is_there():
        wb = Workbook()
        ws = wb.active
        ws.title = "first"
        ws.cell(row=1, column=1, value="name")
        ws.cell(row=1, column=2, value="last date")
        ws.cell(row=1, column=3, value="last payment")
        ws.cell(row=1, column=4, value="total payment")
        wb.save(generate_the_excel_file_name_with_current_year_in_name())

def find_the_next_empty_row(ws):
    row_index = 1
    cell_to_check = ws.cell(row = row_index, column = 1)
    while cell_to_check.value != None:
        row_index = row_index + 1
        cell_to_check = ws.cell(row = row_index, column = 1)
    return row_index

def remove_action_row():
    global row_counter
    if row_counter > 0:
        row_counter -= 1
        dictionary_of_name_entrie[row_counter].destroy()
        dictionary_of_date_entrie[row_counter].destroy()
        dictionary_of_payment_entrie[row_counter].destroy()
        dictionary_of_last_date_labele[row_counter].destroy()
        dictionary_of_last_payment_labele[row_counter].destroy()
        dictionary_of_payment_total_labeles[row_counter].destroy()
        del dictionary_of_last_date_labele_value[row_counter]
        del dictionary_of_last_payment_labele_value[row_counter]
        del dictionary_of_payment_total_labeles_value[row_counter]

def add_action_row():
    global row_counter

    dictionary_of_last_date_labele_value[row_counter] = StringVar()
    dictionary_of_last_payment_labele_value[row_counter] = StringVar()
    dictionary_of_payment_total_labeles_value[row_counter] = StringVar()

    dictionary_of_name_entrie[row_counter] = Entry(main_window_of_gui)
    dictionary_of_name_entrie[row_counter].config(width=20)
    dictionary_of_name_entrie[row_counter].grid(row=row_counter + 1, column = 0)

    dictionary_of_date_entrie[row_counter] = Entry(main_window_of_gui)
    dictionary_of_date_entrie[row_counter].config(width=10)
    dictionary_of_date_entrie[row_counter].grid(row=row_counter + 1, column = 1)

    dictionary_of_payment_entrie[row_counter] = Entry(main_window_of_gui)
    dictionary_of_payment_entrie[row_counter].config(width=10)
    dictionary_of_payment_entrie[row_counter].grid(row=row_counter + 1, column = 2)

    dictionary_of_last_date_labele[row_counter] = Label(main_window_of_gui, textvariable=dictionary_of_last_date_labele_value[row_counter])
    dictionary_of_last_date_labele[row_counter].config(width=10)
    dictionary_of_last_date_labele[row_counter].grid(row=row_counter + 1, column = 3)

    dictionary_of_last_payment_labele[row_counter] = Label(main_window_of_gui, textvariable=dictionary_of_last_payment_labele_value[row_counter])
    dictionary_of_last_payment_labele[row_counter].config(width=10)
    dictionary_of_last_payment_labele[row_counter].grid(row=row_counter + 1, column = 4)

    dictionary_of_payment_total_labeles[row_counter] = Label(main_window_of_gui, textvariable=dictionary_of_payment_total_labeles_value[row_counter])
    dictionary_of_payment_total_labeles[row_counter].config(width=10)
    dictionary_of_payment_total_labeles[row_counter].grid(row=row_counter + 1, column = 5)

    row_counter += 1




def clear_data_from_entries():
    for row in range(row_counter):
        dictionary_of_date_entrie[row].delete(0,'end')
        dictionary_of_payment_entrie[row].delete(0,'end')

def create_missing_sheet(name, excel_workbook):
    print("creating missing sheet " + name)
    if name not in excel_workbook.sheetnames:
        excel_workbook.create_sheet(name)
        ws = excel_workbook[name]
        ws.cell(row=1, column=1, value="last date")
        ws.cell(row=1, column=2, value="last payment")

def update_name_sheet(name, excel_workbook):
    print("Updating " + name)
    ws = excel_workbook[name]
    index = name_index_dictionary[name]
    empty_line_number = find_the_next_empty_row(ws)
    last_date = str(dictionary_of_date_entrie[index].get())
    if last_date != "":
        ws.cell(row=empty_line_number, column=1, value=last_date)
    try:
        last_payment = float(dictionary_of_payment_entrie[index].get())
        ws.cell(row=empty_line_number, column=2, value=last_payment)
    except:
        pass

def update_all_additional_sheets(excel_workbook):
    list_of_names_from_first_sheet = get_list_of_names_from_first_sheet(excel_workbook)
    for name in list_of_names_from_first_sheet:
        create_missing_sheet(name, excel_workbook)
        update_name_sheet(name, excel_workbook)

def update_first_excel_sheet_details(excel_workbook):
    ws = excel_workbook["first"]
    global row_counter
    for index in range(row_counter):
        name = str(dictionary_of_name_entrie[index].get())
        name_index_dictionary[name] = index
        ws.cell(row=index+2, column=1, value=name)
        print(ws.cell(row=2, column=1).value)
        last_date = str(dictionary_of_date_entrie[index].get())
        if last_date != "":
            ws.cell(row=index+2, column=2, value=last_date)
        try:
            last_payment = float(dictionary_of_payment_entrie[index].get())
            ws.cell(row=index+2, column=3, value=last_payment)
            try:
                old_total = float(ws.cell(row=index+2, column=4).value)
            except:
                old_total = 0.0
            new_total = old_total + last_payment
            ws.cell(row=index+2, column=4, value=new_total)
        except:
            pass
        


def get_list_of_names_from_first_sheet(excel_workbook):
    list_of_names = []
    ws = excel_workbook["first"]
    empty_line_number = find_the_next_empty_row(ws)
    for index in range(2, empty_line_number):
        saved_name = str(ws.cell(row=index, column=1).value)
        list_of_names.append(saved_name)
    return list_of_names


def update_first_excel_sheet_and_add_missing_sheets(excel_workbook):
    update_first_excel_sheet_details(excel_workbook)
    update_all_additional_sheets(excel_workbook)



def load_gui_from_excel():
    global row_counter
    while row_counter > 0:
        remove_action_row()
    excel_workbook_name = generate_the_excel_file_name_with_current_year_in_name()
    excel_workbook = load_workbook(excel_workbook_name)
    list_of_names = get_list_of_names_from_first_sheet(excel_workbook)
    for name in list_of_names:
        add_action_row()
    master_dictionary = update_master_dictionary_from_excel(excel_workbook)
    print(master_dictionary)
    for index in range(row_counter):
        saved_name = master_dictionary[index]["name"]
        last_date = master_dictionary[index]["last date"]
        last_payment = master_dictionary[index]["last payment"]
        total_payment = master_dictionary[index]["total payment"]
        dictionary_of_name_entrie[index].insert(0, saved_name)
        dictionary_of_date_entrie[index].delete(0,'end')
        dictionary_of_payment_entrie[index].delete(0,'end')
        dictionary_of_last_date_labele_value[index].set(last_date)
        dictionary_of_last_payment_labele_value[index].set(last_payment)
        dictionary_of_payment_total_labeles_value[index].set(total_payment)

    





def update():
    excel_workbook_name = generate_the_excel_file_name_with_current_year_in_name()
    excel_workbook = load_workbook(excel_workbook_name)
    update_first_excel_sheet_and_add_missing_sheets(excel_workbook)
    excel_workbook.save(excel_workbook_name)
    load_gui_from_excel()

def update_master_dictionary_from_excel(excel_workbook):
    ws = excel_workbook["first"]
    empty_row_number = find_the_next_empty_row(ws)
    for index in range(2, empty_row_number):
        dictionary_index = index - 2
        master_dictionary[dictionary_index] = {}
        saved_name = ws.cell(row=index, column=1).value
        print(saved_name)
        last_date = ws.cell(row=index, column=2).value
        last_payment = ws.cell(row=index, column=3).value
        total_payment = ws.cell(row=index, column=4).value
        master_dictionary[dictionary_index]["name"] = saved_name
        master_dictionary[dictionary_index]["last date"] = last_date
        master_dictionary[dictionary_index]["last payment"] = last_payment
        master_dictionary[dictionary_index]["total payment"] = total_payment
    return master_dictionary




create_excel_file_if_it_was_not_there()

main_window_of_gui = tkinter.Tk()
main_window_of_gui.title("sandbox")
main_window_of_gui.wm_attributes("-topmost", 1)

entry_label_name = Label(main_window_of_gui, text="Name", width=20)
entry_label_name.grid(row=0, column=0)
entry_label_date = Label(main_window_of_gui, text="Date", width=10)
entry_label_date.grid(row=0, column=1)
entry_label_payment = Label(main_window_of_gui, text="Payment", width=10)
entry_label_payment.grid(row=0, column=2)

label_label_last_date = Label(main_window_of_gui, text="Last date", width=10)
label_label_last_date.grid(row=0, column=3)
label_labellast_payment = Label(main_window_of_gui, text="Last payment", width=10)
label_labellast_payment.grid(row=0, column=4)
label_label_total_payment = Label(main_window_of_gui, text="Total payment", width=10)
label_label_total_payment.grid(row=0, column=5)

add_row_button = Button(main_window_of_gui, text="Add", width=10, command=add_action_row)
add_row_button.grid(row=0, column=6)

update_button = Button(main_window_of_gui, text="update", width=10, command=update)
update_button.grid(row=0, column=7)

load_gui_from_excel()

main_window_of_gui.mainloop()