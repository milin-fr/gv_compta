import tkinter
from tkinter import Label, Button, Entry, Text, Checkbutton, OptionMenu, Canvas, Frame, Toplevel, Scrollbar, Listbox, Frame
from tkinter.ttk import Combobox, Treeview
from tkinter import StringVar, IntVar, RIGHT, LEFT, BOTH, Y, END
from tkinter.messagebox import showinfo
import openpyxl
from openpyxl import Workbook, load_workbook
import datetime

'''Below module allows us to interact with Windows files.'''
import os

'''below 3 lines allows script to check the directory where it is executed, so it knows where to crete the excel file. I copied the whole block from stack overflow'''
abspath = os.path.abspath(__file__)
current_directory = os.path.dirname(abspath)
os.chdir(current_directory)

LIST_TYPE_OF_WORK = []
LIST_OF_COMPANIES = []
LIST_OF_BILLS = []
LIST_payment_STATUS = ["Pas payé", "Payé", "Payé", "Annulé"]

class Bill:
    row_placement = ""
    work_type = ""  #1
    company_name = ""  #2
    comment = ""  #3
    start_date = ""  #4
    end_date = ""  #5
    price = ""  #6
    payment_status = ""  #7
    excel_file_name = ""
    def set_excel_name(self):
        self.excel_file_name = "GV compta " + self.work_type + ".xlsx"


def get_file_names_in_script_directory():
    file_names = []
    for root, dirs, files in os.walk(current_directory):
        for filename in files:
            file_names.append(filename)
    return file_names


def get_existing_excel_names():
    '''Each excel file stands for type of work'''
    existing_excel_names = []
    for file_name in get_file_names_in_script_directory():
        if file_name != "GV compta synthese.xlsx" and "GV compta" in file_name:
            existing_excel_names.append(file_name)
    return existing_excel_names


def get_date_dd_mm_yyyy():
    return str(datetime.date.today().strftime("%d.%m.%Y"))


def create_bill_excel_file_if_it_was_not_there(bill_object):
    if bill_object.excel_file_name not in get_existing_excel_names():
        wb = Workbook()
        ws = wb.active
        ws.title = bill_object.company_name
        ws.cell(row=1, column=1, value="work_type")
        ws.cell(row=1, column=2, value="company_name")
        ws.cell(row=1, column=3, value="comment")
        ws.cell(row=1, column=4, value="start_date")
        ws.cell(row=1, column=5, value="end_date")
        ws.cell(row=1, column=6, value="price")
        ws.cell(row=1, column=7, value="payment_status")
        wb.save(bill_object.excel_file_name)


def create_summary_excel_file_if_it_was_not_there():
    if "GV compta synthese.xlsx" not in get_file_names_in_script_directory():
        wb = Workbook()
        ws = wb.active
        ws.title = "Synthese"
        ws.cell(row=1, column=1, value="Budget total")
        ws.cell(row=2, column=1, value="Budget restant estime")
        ws.cell(row=3, column=1, value="Budget restant effectif")
        ws.cell(row=4, column=1, value="Depenses prevues")
        ws.cell(row=5, column=1, value="Depenses effectives")
        wb.create_sheet("Par type de travaux")
        ws = wb["Par type de travaux"]
        ws.cell(row=1, column=1, value="Type de travaux")
        ws.cell(row=1, column=2, value="Depenses prevues")
        ws.cell(row=1, column=3, value="Depenses effectives")
        wb.create_sheet("Par entreprise")
        ws = wb["Par entreprise"]
        ws.cell(row=1, column=1, value="Nom de l'entreprise")
        ws.cell(row=1, column=2, value="Depenses prevues")
        ws.cell(row=1, column=3, value="Depenses effectives")
        wb.save("GV compta synthese.xlsx")
        showinfo("Attention !", "Pour le bon fonctionnement du logiciel, merci de saisir le budget total dans le ficher excel dans la case B1 et relancer le logiciel. Le fichier excel va s'ouvrir automatiquement.")
        os.startfile("GV compta synthese.xlsx")


def update_summary_excel_file():
    get_list_of_bills()
    update_synthese_sheet()
    update_work_type_sheet()
    update_company_sheet()
    update_meta_data_in_root()


def update_synthese_sheet():
    already_spent = 0
    going_to_spend = 0
    for bill in LIST_OF_BILLS:
        if bill.payment_status == "Payé":
            already_spent += float(bill.price)
        if bill.payment_status == "Pas payé":
            going_to_spend += float(bill.price)
    wb = load_workbook("GV compta synthese.xlsx")
    ws = wb["Synthese"]
    try:
        budget = float(ws.cell(row=1, column=2).value)
    except:
        budget = 0
    budget_leftover_estimation = budget - going_to_spend - already_spent
    budget_leftover = budget - already_spent
    ws.cell(row=2, column=2, value=budget_leftover_estimation)
    ws.cell(row=3, column=2, value=budget_leftover)
    ws.cell(row=4, column=2, value=going_to_spend)
    ws.cell(row=5, column=2, value=already_spent)
    wb.save("GV compta synthese.xlsx")


def update_work_type_sheet():
    list_of_work_types = []
    for bill in LIST_OF_BILLS:
        if bill.work_type not in list_of_work_types:
            list_of_work_types.append(bill.work_type)
    wb = load_workbook("GV compta synthese.xlsx")
    ws = wb["Par type de travaux"]
    row_index = 2
    for work_type in list_of_work_types:
        already_spent = 0
        going_to_spend = 0
        for bill in LIST_OF_BILLS:
            if bill.work_type == work_type:
                if bill.payment_status == "Payé":
                    already_spent += float(bill.price)
                if bill.payment_status == "Pas payé":
                    going_to_spend += float(bill.price)
                ws.cell(row=row_index, column=1, value=bill.work_type)
                ws.cell(row=row_index, column=2, value=going_to_spend)
                ws.cell(row=row_index, column=3, value=already_spent)
        row_index +=1
    wb.save("GV compta synthese.xlsx")


def update_company_sheet():
    list_of_companies = []
    for bill in LIST_OF_BILLS:
        if bill.company_name not in list_of_companies:
            list_of_companies.append(bill.company_name)
    wb = load_workbook("GV compta synthese.xlsx")
    ws = wb["Par entreprise"]
    row_index = 2
    for company_name in list_of_companies:
        already_spent = 0
        going_to_spend = 0
        for bill in LIST_OF_BILLS:
            if bill.company_name == company_name:
                if bill.payment_status == "Payé":
                    already_spent += float(bill.price)
                if bill.payment_status == "Pas payé":
                    going_to_spend += float(bill.price)
                ws.cell(row=row_index, column=1, value=bill.company_name)
                ws.cell(row=row_index, column=2, value=going_to_spend)
                ws.cell(row=row_index, column=3, value=already_spent)
        row_index +=1
    wb.save("GV compta synthese.xlsx")


def find_the_next_empty_row(ws):
    row_index = 1
    cell_to_check = ws.cell(row = row_index, column = 1)
    while cell_to_check.value != None:
        row_index = row_index + 1
        cell_to_check = ws.cell(row = row_index, column = 1)
    return row_index


def create_missing_sheet_if_it_was_not_there(bill_object):
    wb = load_workbook(bill_object.excel_file_name)
    if bill_object.company_name not in wb.sheetnames:
        wb.create_sheet(bill_object.company_name)
        ws = wb[bill_object.company_name]
        ws.cell(row=1, column=1, value="work_type")
        ws.cell(row=1, column=2, value="company_name")
        ws.cell(row=1, column=3, value="comment")
        ws.cell(row=1, column=4, value="start_date")
        ws.cell(row=1, column=5, value="end_date")
        ws.cell(row=1, column=6, value="price")
        ws.cell(row=1, column=7, value="payment_status")
        wb.save(bill_object.excel_file_name)
    else:
        wb.close()


def toplevel_was_closed(evt):
    unblock_root_buttons()


def new_bill():
    bloc_root_buttons()
    today_date_yyyy_mm_dd = get_date_dd_mm_yyyy()
    update_work_type_list()
    window_new_bill = Toplevel()
    x = main_window_of_gui.winfo_x()
    y = main_window_of_gui.winfo_y()
    w = main_window_of_gui.winfo_width()
    h = main_window_of_gui.winfo_height()
    window_new_bill.geometry("%dx%d+%d+%d" % (w, h, x, y))
    window_new_bill.title("Informations additionelles")
    window_new_bill.wm_attributes("-topmost", 1)
    label_work_selection = Label(window_new_bill, text = "Type de travaux :", width=15)
    label_work_selection.grid(column=0, row=0, pady=5)
    combo_work_selection = Combobox(window_new_bill, values = LIST_TYPE_OF_WORK)
    combo_work_selection.grid(column=1, row=0, columnspan=2, pady=5)
    combo_work_selection.bind("<<ComboboxSelected>>", lambda evt: update_company_name_list(combo_work_selection, combo_company_selection))
    combo_work_selection.bind("<Return>", lambda evt: update_company_name_list(combo_work_selection, combo_company_selection))
    label_company_selection = Label(window_new_bill, text = "Nom de l'entreprise :", width=15)
    label_company_selection.grid(column=0, row=1, pady=5)
    combo_company_selection = Combobox(window_new_bill, values = LIST_OF_COMPANIES)
    combo_company_selection.grid(column=1, row=1, columnspan=2, pady=5)
    label_price = Label(window_new_bill, text = "Prix :", width=15)
    label_price.grid(column=0, row=2, pady=5)
    entry_price = Entry(window_new_bill, width=23)
    entry_price.grid(column=1, row=2, columnspan=2, pady=5)
    label_start_date = Label(window_new_bill, text = "Date de debut :", width=15)
    label_start_date.grid(column=0, row=3, pady=5)
    entry_start_date = Entry(window_new_bill, width=23)
    entry_start_date.insert(0, today_date_yyyy_mm_dd)
    entry_start_date.grid(column=1, row=3, columnspan=2, pady=5)
    label_end_date = Label(window_new_bill, text = "Date de fin :", width=15)
    label_end_date.grid(column=0, row=4, pady=5)
    entry_end_date = Entry(window_new_bill, width=23)
    entry_end_date.insert(0, today_date_yyyy_mm_dd)
    entry_end_date.grid(column=1, row=4, columnspan=2, pady=5)
    label_status = Label(window_new_bill, text = "Etat du payment :", width=15)
    label_status.grid(column=0, row=5, pady=5)
    var_payment_status = StringVar()
    var_payment_status.set(LIST_payment_STATUS[0])
    dropdown_payment_status = OptionMenu(window_new_bill, var_payment_status, *LIST_payment_STATUS)
    dropdown_payment_status.grid(column=1, row=5, columnspan=2, pady=5)
    dropdown_payment_status.config(width=18)
    label_comment = Label(window_new_bill, text = "Commentaires :", width=15)
    label_comment.grid(column=0, row=6, columnspan=3, pady=5)
    text_comment = Text(window_new_bill, width=60, height=10)
    text_comment.grid(column=0, row=7, columnspan=3, pady=5)
    data_entries = [combo_work_selection, combo_company_selection, text_comment, entry_start_date, entry_end_date, entry_price, var_payment_status]
    button_confirm_new_bill = Button(window_new_bill, text="Confirmer", width=10, height=3, command=lambda: confirm_new_bill(data_entries, window_new_bill))
    button_confirm_new_bill.grid(column=0, row=8, pady=5)
    button_cancel_new_bill = Button(window_new_bill, text="Annuler", width=10, height=3, command=lambda: cancel_current_window(window_new_bill))
    button_cancel_new_bill.grid(column=2, row=8, pady=5)
    button_cancel_new_bill.bind("<Destroy>", toplevel_was_closed)  # if bind on toplevel, the destruction of all widgets in toplevel trigers the function


def existing_bill(bill_to_edit):
    bloc_root_buttons()
    window_bill_update = Toplevel()
    x = main_window_of_gui.winfo_x()
    y = main_window_of_gui.winfo_y()
    w = main_window_of_gui.winfo_width()
    h = main_window_of_gui.winfo_height()
    window_bill_update.geometry("%dx%d+%d+%d" % (w, h, x, y))
    window_bill_update.title("Editer une facture existente")
    window_bill_update.wm_attributes("-topmost", 1)
    label_work_selection = Label(window_bill_update, text = "Type de travaux :", width=15)
    label_work_selection.grid(column=0, row=0, pady=5)
    combo_work_selection = Combobox(window_bill_update, values = LIST_TYPE_OF_WORK)
    combo_work_selection.grid(column=1, row=0, columnspan=2, pady=5)
    combo_work_selection.set(bill_to_edit.work_type)
    combo_work_selection.config(state="disabled")
    label_company_selection = Label(window_bill_update, text = "Nom de l'entreprise :", width=15)
    label_company_selection.grid(column=0, row=1, pady=5)
    combo_company_selection = Combobox(window_bill_update, values = LIST_OF_COMPANIES)
    combo_company_selection.grid(column=1, row=1, columnspan=2, pady=5)
    combo_company_selection.set(bill_to_edit.company_name)
    combo_company_selection.config(state="disabled")
    label_price = Label(window_bill_update, text = "Prix :", width=15)
    label_price.grid(column=0, row=2, pady=5)
    entry_price = Entry(window_bill_update, width=23)
    entry_price.grid(column=1, row=2, columnspan=2, pady=5)
    entry_price.insert(0, bill_to_edit.price)
    label_start_date = Label(window_bill_update, text = "Date de debut :", width=15)
    label_start_date.grid(column=0, row=3, pady=5)
    entry_start_date = Entry(window_bill_update, width=23)
    entry_start_date.insert(0, bill_to_edit.start_date)
    entry_start_date.grid(column=1, row=3, columnspan=2, pady=5)
    label_end_date = Label(window_bill_update, text = "Date de fin :", width=15)
    label_end_date.grid(column=0, row=4, pady=5)
    entry_end_date = Entry(window_bill_update, width=23)
    entry_end_date.insert(0, bill_to_edit.end_date)
    entry_end_date.grid(column=1, row=4, columnspan=2, pady=5)
    label_status = Label(window_bill_update, text = "Etat du payment :", width=15)
    label_status.grid(column=0, row=5, pady=5)
    var_payment_status = StringVar()
    var_payment_status.set(bill_to_edit.payment_status)
    dropdown_payment_status = OptionMenu(window_bill_update, var_payment_status, *LIST_payment_STATUS)
    dropdown_payment_status.grid(column=1, row=5, columnspan=2, pady=5)
    dropdown_payment_status.config(width=18)
    label_comment = Label(window_bill_update, text = "Commentaires :", width=15)
    label_comment.grid(column=3, row=0, pady=5)
    text_comment = Text(window_bill_update, width=60, height=10)
    text_comment.grid(column=3, row=1, rowspan=5, pady=5)
    text_comment.insert(END, bill_to_edit.comment)
    row_placement = bill_to_edit.row_placement
    data_entries = [combo_work_selection, combo_company_selection, text_comment, entry_start_date, entry_end_date, entry_price, var_payment_status, row_placement]
    button_confirm_bill_update = Button(window_bill_update, text="Confirmer", width=10, height=3, command=lambda: confirm_bill_update(data_entries, window_bill_update))
    button_confirm_bill_update.grid(column=0, row=8, pady=5)
    button_cancel_bill_update = Button(window_bill_update, text="Annuler", width=10, height=3, command=lambda: cancel_current_window(window_bill_update))
    button_cancel_bill_update.grid(column=2, row=8, pady=5)
    button_cancel_bill_update.bind("<Destroy>", toplevel_was_closed)  # if bind on toplevel, the destruction of all widgets in toplevel trigers the function


def cancel_current_window(window_to_close):
    window_to_close.destroy()


def confirm_new_bill(data_entries, window_to_close):
    bill_object = Bill()
    bill_object.work_type = data_entries[0].get()
    bill_object.company_name = data_entries[1].get()
    bill_object.comment = data_entries[2].get('1.0', 'end-1c')
    bill_object.start_date = data_entries[3].get()
    bill_object.end_date = data_entries[4].get()
    bill_object.price = data_entries[5].get().replace(",", ".")
    bill_object.payment_status = data_entries[6].get()
    bill_object.set_excel_name()
    create_bill_excel_file_if_it_was_not_there(bill_object)
    create_missing_sheet_if_it_was_not_there(bill_object)
    save_bill_in_excel(bill_object)
    cancel_current_window(window_to_close)


def confirm_bill_update(data_entries, window_to_close):
    bill_object = Bill()
    bill_object.work_type = data_entries[0].get()
    bill_object.company_name = data_entries[1].get()
    bill_object.comment = data_entries[2].get('1.0', 'end-1c')
    bill_object.start_date = data_entries[3].get()
    bill_object.end_date = data_entries[4].get()
    bill_object.price = data_entries[5].get()
    bill_object.payment_status = data_entries[6].get()
    bill_object.row_placement = data_entries[7]
    bill_object.set_excel_name()
    create_bill_excel_file_if_it_was_not_there(bill_object)
    create_missing_sheet_if_it_was_not_there(bill_object)
    save_bill_in_excel(bill_object)
    cancel_current_window(window_to_close)


def save_bill_in_excel(bill_object):
    wb = load_workbook(bill_object.excel_file_name)
    ws = wb[bill_object.company_name]
    if bill_object.row_placement == "":
        row_of_this_bill = find_the_next_empty_row(ws)
    else:
        row_of_this_bill = int(bill_object.row_placement)
    ws.cell(row=row_of_this_bill, column=1, value=bill_object.work_type)
    ws.cell(row=row_of_this_bill, column=2, value=bill_object.company_name)
    ws.cell(row=row_of_this_bill, column=3, value=bill_object.comment)
    ws.cell(row=row_of_this_bill, column=4, value=bill_object.start_date)
    ws.cell(row=row_of_this_bill, column=5, value=bill_object.end_date)
    ws.cell(row=row_of_this_bill, column=6, value=bill_object.price)
    ws.cell(row=row_of_this_bill, column=7, value=bill_object.payment_status)
    wb.save(bill_object.excel_file_name)
    update_summary_excel_file()


def get_list_of_names_from_first_sheet(excel_workbook):
    list_of_names = []
    ws = excel_workbook["first"]
    empty_line_number = find_the_next_empty_row(ws)
    for index in range(2, empty_line_number):
        saved_name = str(ws.cell(row=index, column=1).value)
        list_of_names.append(saved_name)
    return list_of_names


def update_work_type_list():
    global LIST_TYPE_OF_WORK
    LIST_TYPE_OF_WORK = []
    existing_excel_names = get_existing_excel_names()
    for file_name in existing_excel_names:
        work_type = file_name[10:-5]
        LIST_TYPE_OF_WORK.append(work_type)


def update_company_name_list(combo_work_selection, combo_company_selection):
    global LIST_OF_COMPANIES
    LIST_OF_COMPANIES = []
    excel_file_name = "GV compta " + combo_work_selection.get() + ".xlsx"
    if excel_file_name in get_existing_excel_names():
        wb = load_workbook(excel_file_name)
        LIST_OF_COMPANIES = wb.sheetnames
        wb.close()
        combo_company_selection['values'] = LIST_OF_COMPANIES
        combo_company_selection.current(0)
    combo_company_selection.focus_set()


def doube_click_bill_line(evt, window_to_close):
    clicked_widger = evt.widget
    row_id = clicked_widger.selection()[0] #particular line that is selected
    text = clicked_widger.item(row_id, 'text')
    bill_to_edit = Bill()
    bill_to_edit.row_placement = clicked_widger.item(row_id, 'text')
    bill_to_edit.work_type = clicked_widger.item(row_id, 'values')[0]
    bill_to_edit.company_name = clicked_widger.item(row_id, 'values')[1]
    bill_to_edit.comment = clicked_widger.item(row_id, 'values')[2]
    bill_to_edit.start_date = clicked_widger.item(row_id, 'values')[3]
    bill_to_edit.end_date = clicked_widger.item(row_id, 'values')[4]
    bill_to_edit.price = clicked_widger.item(row_id, 'values')[5]
    bill_to_edit.payment_status = clicked_widger.item(row_id, 'values')[6]
    bill_to_edit.set_excel_name()
    existing_bill(bill_to_edit)
    cancel_current_window(window_to_close)


def treeview_sort_column(tv, col, reverse):
    list_of_lines = tv.get_children('')
    list_of_something = []
    for single_line in list_of_lines:
        list_of_something.append((tv.set(single_line, col), single_line))
    list_of_something.sort(reverse=reverse)
    for index, (val, single_line) in enumerate(list_of_something):
        tv.move(single_line, '', index)
    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))


def open_ongoing_view():
    bloc_root_buttons()
    get_list_of_bills()
    window_ongoing_bill = Toplevel()
    x = main_window_of_gui.winfo_x()
    y = main_window_of_gui.winfo_y()
    w = main_window_of_gui.winfo_width()
    h = main_window_of_gui.winfo_height()
    window_ongoing_bill.geometry("%dx%d+%d+%d" % (w, h, x, y))
    window_ongoing_bill.title("Factures en cours")
    window_ongoing_bill.wm_attributes("-topmost", 1)
    frame_for_the_list = Frame(window_ongoing_bill)
    frame_for_the_list.grid(column=0, row=0)
    tv_columns = ('work_type', 'company_name', 'comment', "start_date", "end_date", "price", "payment_status")
    treeview_details_of_ongoing_bills = Treeview(frame_for_the_list, columns=tv_columns, show='headings')
    for column in tv_columns:
        treeview_details_of_ongoing_bills.heading(column, text=column, command=lambda col=column: treeview_sort_column(treeview_details_of_ongoing_bills, col, False))
        treeview_details_of_ongoing_bills.column(column, anchor='center', width=100)
    treeview_details_of_ongoing_bills.column('comment', anchor='center', width=300)
    scrollbar = Scrollbar(frame_for_the_list, command=treeview_details_of_ongoing_bills.yview)
    scrollbar.pack(side=RIGHT, fill=Y)
    treeview_details_of_ongoing_bills.pack()
    for bill in LIST_OF_BILLS:
        treeview_details_of_ongoing_bills.insert('', 'end', text=str(bill.row_placement), 
                            values=(
                                bill.work_type,
                                bill.company_name,
                                bill.comment,
                                bill.start_date,
                                bill.end_date,
                                bill.price,
                                bill.payment_status
                            ))
    treeview_details_of_ongoing_bills.bind('<Double-1>', lambda event: doube_click_bill_line(event, window_ongoing_bill))
    treeview_details_of_ongoing_bills.configure(yscrollcommand=scrollbar.set)
    label_explanation = Label(window_ongoing_bill, text="Double-click pour editer une facture.")
    label_explanation.grid(column=0, row=1)
    treeview_details_of_ongoing_bills.bind("<Destroy>", toplevel_was_closed)  # if bind on toplevel, the destruction of all widgets in toplevel trigers the function


def get_details_out_of_excel(excel_file):
    global LIST_OF_BILLS
    wb = load_workbook(excel_file)
    for company_sheet in wb.sheetnames:
        ws = wb[company_sheet]
        maximum_number_of_rows = find_the_next_empty_row(ws)
        for current_row in range(2, maximum_number_of_rows):
            bill = Bill()
            bill.row_placement = current_row
            bill.work_type = ws.cell(row=current_row, column=1).value
            bill.company_name = ws.cell(row=current_row, column=2).value
            bill.comment = ws.cell(row=current_row, column=3).value
            bill.start_date = ws.cell(row=current_row, column=4).value
            bill.end_date = ws.cell(row=current_row, column=5).value
            bill.price = ws.cell(row=current_row, column=6).value
            bill.payment_status = ws.cell(row=current_row, column=7).value
            bill.set_excel_name()
            LIST_OF_BILLS.append(bill)
    wb.close()


def get_list_of_bills():
    global LIST_OF_BILLS
    LIST_OF_BILLS = []
    existing_excel_names = get_existing_excel_names()
    for excel_file in existing_excel_names:
        get_details_out_of_excel(excel_file)


def bloc_root_buttons():
    button_start_new_work.config(state="disabled")
    button_view_ongoing_work.config(state="disabled")


def unblock_root_buttons():
    button_start_new_work.config(state="normal")
    button_view_ongoing_work.config(state="normal")


def update_meta_data_in_root():
    create_global_meta_treeview()
    update_work_type_meta_treeview()
    update_company_meta_treeview()


def create_global_meta_treeview():
    var_budget = StringVar()
    var_budget_leftover_estimation = StringVar()
    var_budget_leftover = StringVar()
    var_will_spend = StringVar()
    var_already_spent = StringVar()
    wb = load_workbook("GV compta synthese.xlsx")
    ws = wb["Synthese"]
    var_budget.set(str(round(ws.cell(row=1, column=2).value, 2)))
    var_budget_leftover_estimation.set(str(round(ws.cell(row=2, column=2).value, 2)))
    var_budget_leftover.set(str(round(ws.cell(row=3, column=2).value, 2)))
    var_will_spend.set(str(round(ws.cell(row=4, column=2).value, 2)))
    var_already_spent.set(str(round(ws.cell(row=5, column=2).value, 2)))
    wb.close()
    label_budget_intro = Label(frame_global_meta_tree_view, text="Budget total :")
    label_budget_intro.grid(column=0, row=0, sticky="w")
    label_budget = Label(frame_global_meta_tree_view, textvariable=var_budget)
    label_budget.grid(column=1, row=0, sticky="e")
    label_budget_leftover_estimation_intro = Label(frame_global_meta_tree_view, text="Budget restant previsionel :")
    label_budget_leftover_estimation_intro.grid(column=0, row=1, sticky="w")
    label_budget_leftover_estimation = Label(frame_global_meta_tree_view, textvariable=var_budget_leftover_estimation)
    label_budget_leftover_estimation.grid(column=1, row=1, sticky="e")
    label_budget_leftover_intro = Label(frame_global_meta_tree_view, text="Budget restant effectif :")
    label_budget_leftover_intro.grid(column=0, row=2, sticky="w")
    label_budget_leftover = Label(frame_global_meta_tree_view, textvariable=var_budget_leftover)
    label_budget_leftover.grid(column=1, row=2, sticky="e")
    label_will_to_spend_intro = Label(frame_global_meta_tree_view, text="Depenses prevues :")
    label_will_to_spend_intro.grid(column=0, row=3, sticky="w")
    label_will_to_spend = Label(frame_global_meta_tree_view, textvariable=var_will_spend)
    label_will_to_spend.grid(column=1, row=3, sticky="e")
    label_already_spent_intro = Label(frame_global_meta_tree_view, text="Depenses effectives :")
    label_already_spent_intro.grid(column=0, row=4, sticky="w")
    label_already_spent = Label(frame_global_meta_tree_view, textvariable=var_already_spent)
    label_already_spent.grid(column=1, row=4, sticky="e")


def update_work_type_meta_treeview():
    tv_work_type.delete(*tv_work_type.get_children())
    wb = load_workbook("GV compta synthese.xlsx")
    ws = wb["Par type de travaux"]
    empty_row = find_the_next_empty_row(ws)
    for row in range(2, empty_row):
        tv_work_type.insert('', 'end', text=str(row),
                                values=(
                                    str(ws.cell(row=row, column=1).value),
                                    str(ws.cell(row=row, column=2).value),
                                    str(ws.cell(row=row, column=3).value)
                                ))
    wb.close()


def update_company_meta_treeview():
    tv_company_name.delete(*tv_company_name.get_children())
    wb = load_workbook("GV compta synthese.xlsx")
    ws = wb["Par entreprise"]
    empty_row = find_the_next_empty_row(ws)
    for row in range(2, empty_row):
        tv_company_name.insert('', 'end', text=str(row),
                                values=(
                                    str(ws.cell(row=row, column=1).value),
                                    str(ws.cell(row=row, column=2).value),
                                    str(ws.cell(row=row, column=3).value)
                                ))
    wb.close()
    

create_summary_excel_file_if_it_was_not_there()

main_window_of_gui = tkinter.Tk()
main_window_of_gui.title("sandbox")
main_window_of_gui.geometry("1000x600")
main_window_of_gui.resizable(0, 0)

label_work_type_meta_tree_view = Label(main_window_of_gui, text="Donnees par type de travaux :")
label_work_type_meta_tree_view.grid(column=0, row=0, columnspan=2)

frame_work_type_meta_tree_view = Frame(main_window_of_gui)
frame_work_type_meta_tree_view.grid(column=0, row=1, columnspan=2)

tv_work_type_colums = ("Type de travaux", "Depenses prevues", "Depenses effectives")
tv_work_type = Treeview(frame_work_type_meta_tree_view, columns=tv_work_type_colums, show='headings')
for column in tv_work_type_colums:
    tv_work_type.heading(column, text=column, command=lambda col=column: treeview_sort_column(tv_work_type, col, False))
    tv_work_type.column(column, anchor='center', width=150)
scrollbar = Scrollbar(frame_work_type_meta_tree_view, command=tv_work_type.yview)
scrollbar.pack(side=RIGHT, fill=Y)
tv_work_type.pack()

label_company_meta_tree_view = Label(main_window_of_gui, text="Donnees par entreprise :")
label_company_meta_tree_view.grid(column=2, row=0, columnspan=2)

frame_company_meta_tree_view = Frame(main_window_of_gui)
frame_company_meta_tree_view.grid(column=2, row=1, columnspan=2)

tv_company_name_colums = ("Nom de l'entreprise", "Depenses prevues", "Depenses effectives")
tv_company_name = Treeview(frame_company_meta_tree_view, columns=tv_company_name_colums, show='headings')
for column in tv_company_name_colums:
    tv_company_name.heading(column, text=column, command=lambda col=column: treeview_sort_column(tv_company_name, col, False))
    tv_company_name.column(column, anchor='center', width=150)
scrollbar = Scrollbar(frame_company_meta_tree_view, command=tv_company_name.yview)
scrollbar.pack(side=RIGHT, fill=Y)
tv_company_name.pack()

label_global_meta_tree_view = Label(main_window_of_gui, text="Donnees d'ensemble :")
label_global_meta_tree_view.grid(column=0, row=3, columnspan=2)

frame_global_meta_tree_view = Frame(main_window_of_gui)
frame_global_meta_tree_view.grid(column=0, row=4, columnspan=2, sticky="n")

button_start_new_work = Button(main_window_of_gui, text="Nouvelle facture", width=20, height=3, command=new_bill)
button_start_new_work.grid(column=2, row=3)

button_view_ongoing_work = Button(main_window_of_gui, text="Facture en cours", width=20, height=3, command=open_ongoing_view)
button_view_ongoing_work.grid(column=2, row=4)

update_summary_excel_file()

main_window_of_gui.mainloop()
